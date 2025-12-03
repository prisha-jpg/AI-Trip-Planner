"""Utility functions for exporting and data processing."""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

def clean_data_for_excel(data_items):
    """Clean data items by converting lists to comma-separated strings."""
    cleaned_items = []
    for item in data_items:
        cleaned_item = {}
        for key, value in item.items():
            if isinstance(value, list):
                cleaned_item[key] = ", ".join(str(v) for v in value)
            elif isinstance(value, dict):
                cleaned_item[key] = str(value)
            else:
                cleaned_item[key] = value
        cleaned_items.append(cleaned_item)
    return cleaned_items

def export_to_excel(state, filename: str):
    """Export trip plan to Excel"""
    try:
        wb = openpyxl.Workbook()
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Trip Summary
        ws = wb.create_sheet("Trip Summary")
        ws['A1'] = f"Trip Plan: {state.get('from_city', 'N/A')} → {state.get('to_city', 'N/A')}"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:D1')
        
        summary_data = [
            ["From", state.get('from_city', 'N/A')],
            ["To", state.get('to_city', 'N/A')],
            ["Arrival Date", state.get('arrival_date', 'N/A')],
            ["Duration", f"{state.get('num_days', 0)} days"],
            ["Travelers", f"{state.get('num_adults', 0)} adults, {state.get('num_kids', 0)} children"]
        ]
        
        for idx, row in enumerate(summary_data, start=3):
            ws[f'A{idx}'] = row[0]
            ws[f'B{idx}'] = row[1]
            ws[f'A{idx}'].font = Font(bold=True)
        
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 30
        
        # Weather Sheet
        if state.get("weather_data") and "forecasts" in state["weather_data"]:
            weather_ws = wb.create_sheet("Weather Forecast")
            weather_df = pd.DataFrame(state["weather_data"]["forecasts"])
            
            # Add headers
            weather_ws['A1'] = "5-Day Weather Forecast"
            weather_ws['A1'].font = Font(size=14, bold=True)
            weather_ws.merge_cells('A1:G1')
            
            # Add data
            for r in dataframe_to_rows(weather_df, index=False, header=True):
                weather_ws.append(r)
            
            # Format headers
            for cell in weather_ws[3]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Attractions Sheet
        if state.get("attractions_data") and "items" in state["attractions_data"]:
            attractions_ws = wb.create_sheet("Top Attractions")
            # Clean attractions data - convert any lists to strings
            attractions_items = clean_data_for_excel(state["attractions_data"]["items"])
            attractions_df = pd.DataFrame(attractions_items)
            
            attractions_ws['A1'] = "Top Attractions"
            attractions_ws['A1'].font = Font(size=14, bold=True)
            attractions_ws.merge_cells('A1:H1')
            
            for r in dataframe_to_rows(attractions_df, index=False, header=True):
                attractions_ws.append(r)
            
            for cell in attractions_ws[3]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Hotels Sheet
        if state.get("hotel_data") and "items" in state["hotel_data"]:
            hotels_ws = wb.create_sheet("Hotel Recommendations")
            # Clean hotel data - convert any lists (like amenities) to strings
            hotel_items = clean_data_for_excel(state["hotel_data"]["items"])
            hotels_df = pd.DataFrame(hotel_items)
            
            hotels_ws['A1'] = "Hotel Recommendations"
            hotels_ws['A1'].font = Font(size=14, bold=True)
            hotels_ws.merge_cells('A1:H1')
            
            for r in dataframe_to_rows(hotels_df, index=False, header=True):
                hotels_ws.append(r)
            
            for cell in hotels_ws[3]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Expenses Sheet
        if state.get("expenses_data") and "items" in state["expenses_data"]:
            expenses_ws = wb.create_sheet("Trip Expenses")
            # Clean expenses data - convert any lists to strings
            expense_items = clean_data_for_excel(state["expenses_data"]["items"])
            expenses_df = pd.DataFrame(expense_items)
            
            expenses_ws['A1'] = "Trip Expenses Breakdown"
            expenses_ws['A1'].font = Font(size=14, bold=True)
            expenses_ws.merge_cells('A1:D1')
            
            for r in dataframe_to_rows(expenses_df, index=False, header=True):
                expenses_ws.append(r)
            
            for cell in expenses_ws[3]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            # Highlight total row
            total_row = len(expenses_df) + 3
            for cell in expenses_ws[total_row]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        
        # Itinerary Sheet
        if state.get("itinerary"):
            itinerary_ws = wb.create_sheet("Detailed Itinerary")
            itinerary_ws['A1'] = "Detailed Day-by-Day Itinerary"
            itinerary_ws['A1'].font = Font(size=14, bold=True)
            
            # Split itinerary into lines and add to sheet
            lines = state["itinerary"].split('\n')
            for idx, line in enumerate(lines, start=3):
                itinerary_ws[f'A{idx}'] = line
            
            itinerary_ws.column_dimensions['A'].width = 100
        
        wb.save(filename)
        return filename
        
    except Exception as e:
        # Log the error and re-raise with more context
        error_msg = f"Error exporting to Excel: {str(e)}"
        raise ValueError(error_msg) from e